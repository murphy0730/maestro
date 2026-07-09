from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
FONT = "Arial"

# ---------- styles ----------
title_font = Font(name=FONT, size=14, bold=True, color="1F3864")
hdr_font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
cell_font = Font(name=FONT, size=10, color="000000")
bold_cell = Font(name=FONT, size=10, bold=True, color="000000")

hdr_fill = PatternFill("solid", fgColor="1F4E79")
green_fill = PatternFill("solid", fgColor="C6EFCE")   # 已覆盖/增强
yellow_fill = PatternFill("solid", fgColor="FFEB9C")  # 部分覆盖
red_fill = PatternFill("solid", fgColor="FFC7CE")      # 缺失
band_fill = PatternFill("solid", fgColor="F2F6FC")

status_map = {
    "已覆盖": green_fill,
    "增强(我独有)": green_fill,
    "部分覆盖": yellow_fill,
    "缺失": red_fill,
}
status_font = {
    "已覆盖": Font(name=FONT, size=10, bold=True, color="006100"),
    "增强(我独有)": Font(name=FONT, size=10, bold=True, color="006100"),
    "部分覆盖": Font(name=FONT, size=10, bold=True, color="9C6500"),
    "缺失": Font(name=FONT, size=10, bold=True, color="9C0006"),
}

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ================= Sheet 1: 能力对比总表 =================
s1 = wb.active
s1.title = "能力对比总表"

s1["A1"] = "工具调用链能力对比：Claude Code vs 我的实现"
s1["A1"].font = title_font
s1.merge_cells("A1:G1")
s1["A2"] = "标注规则：✅已覆盖=等效实现；🟢增强=我方独有或更强；🟡部分覆盖=仅部分实现；🔴缺失=链路上未体现"
s1["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
s1.merge_cells("A2:G2")

# summary counts via COUNTIF
s1["A4"] = "统计"
s1["A4"].font = bold_cell
s1["B4"] = "已覆盖"; s1["C4"] = "增强(我独有)"; s1["D4"] = "部分覆盖"; s1["E4"] = "缺失"; s1["F4"] = "合计"
for c in ("B4","C4","D4","E4","F4"):
    s1[c].font = hdr_font; s1[c].fill = hdr_fill; s1[c].alignment = center
s1["B5"] = "=COUNTIF(G8:G26,B4)"
s1["C5"] = "=COUNTIF(G8:G26,C4)"
s1["D5"] = "=COUNTIF(G8:G26,D4)"
s1["E5"] = "=COUNTIF(G8:G26,E4)"
s1["F5"] = "=SUM(B5:E5)"
for c in ("B5","C5","D5","E5","F5"):
    s1[c].font = bold_cell; s1[c].alignment = center; s1[c].border = border

headers = ["#", "能力分类", "能力项", "Claude Code 实现环节", "我的实现环节", "差距/说明", "覆盖状态"]
hrow = 7
for i, h in enumerate(headers, start=1):
    c = s1.cell(row=hrow, column=i, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

rows = [
    # id, category, capability, claude step, mine step, gap note, status
    (1, "解析与编排", "解析 tool_use（name + input）",
     "① API 返回 tool_use block",
     "① llm.chat_turn() 返回 tool_calls[]（OpenAI 格式，支持一轮多个 call）",
     "等效且更强：我方原生支持一轮多 call 并行返回", "增强(我独有)"),
    (2, "解析与编排", "多工具批量/并发执行",
     "② StreamingToolExecutor.runTools()",
     "③ 逐 tool_call → _handle_call()，串行处理",
     "仅串行，未见 runTools 式并发；多 call 是顺序执行而非并发", "部分覆盖"),
    (3, "解析与编排", "外层循环 + 步数硬上限",
     "（隐式，无显式 max_steps）",
     "② run() 硬上限 max_steps 防失控",
     "我方独有：显式步数上限", "增强(我独有)"),
    (4, "工具查找", "findToolByName 工具定位",
     "③ findToolByName()",
     "⑥ self._tools.get(name)",
     "等效实现", "已覆盖"),
    (5, "输入校验", "泛型输入校验 validateInput",
     "④ validateInput()（失败返回错误 tool_result）",
     "⑦ 仅写操作有 precondition(args) 前置断言",
     "缺失：读/中性工具无统一 schema 校验层，仅写操作有断言", "缺失"),
    (6, "输入校验", "技能级追加断言",
     "（无对应）",
     "⑦b self._extra 技能级追加断言",
     "我方独有：技能维度的二次校验", "增强(我独有)"),
    (7, "权限控制", "交互式权限确认 canUseTool（Ask UI）",
     "⑤ canUseTool() Ask 模式弹确认（拒绝→拒绝 tool_result）",
     "⑧ 写操作走 ActionGate.request（auto/pending/deny）",
     "部分覆盖：仅写操作有 pending 询问；读/中性工具默认放行，无通用'询问用户确认'交互层", "部分覆盖"),
    (8, "权限控制", "规则匹配权限 checkPermissions",
     "⑥ checkPermissions() 规则引擎",
     "⑧ ActionGate 规则，但与写操作耦合",
     "部分覆盖：无独立的通用权限规则引擎，规则散落在写操作授权中", "部分覆盖"),
    (9, "权限控制", "工具白名单",
     "（部分由 checkPermissions 覆盖）",
     "④ 护栏3 name not in self._allowed 直接拦截",
     "我方独有：显式工具白名单前置拦截", "增强(我独有)"),
    (10, "执行", "实际执行 call()",
     "⑦ call() → 实际操作",
     "⑧ self._tools.execute(name,args) → handler(**args)",
     "等效实现", "已覆盖"),
    (11, "进度反馈", "实时进度回调 onProgress",
     "⑦ call() 内 onProgress() 实时更新 UI",
     "（链路未体现，handler 执行无流式进度回调）",
     "缺失：长任务执行期间 UI 无实时进度反馈", "缺失"),
    (12, "结果序列化", "结果→API 格式映射",
     "⑨ mapToolResultToToolResultBlockParam()",
     "⑩ _serialize_observation() + ⑪ append {role:tool,...}",
     "等效实现", "已覆盖"),
    (13, "结果序列化", "超大结果截断",
     "（隐式处理）",
     "⑩ _serialize_observation() >8KB 截断",
     "我方独有：显式 8KB 截断护栏", "增强(我独有)"),
    (14, "结果序列化", "结果对象封装 ToolResult<Output>",
     "⑧ 返回 ToolResult<Output>",
     "⑧ handler 返回 + _serialize_observation 封装",
     "等效实现", "已覆盖"),
    (15, "护栏/防错", "绕圈去重",
     "（无对应）",
     "⑤ seen[(name,args)]>1 跳过",
     "我方独有：防重复调用死循环", "增强(我独有)"),
    (16, "护栏/防错", "卡死软检测",
     "（无对应）",
     "② _is_stuck 软检测",
     "我方独有：卡死自动干预", "增强(我独有)"),
    (17, "护栏/防错", "空响应 nudge",
     "（无显式）",
     "③ EMPTY 分支 → nudge 引导",
     "我方独有：空响应主动引导", "增强(我独有)"),
    (18, "护栏/防错", "写后清读 seen 计数",
     "（无对应）",
     "⑨ 写操作后清读类 seen（状态已变，允许重读）",
     "我方独有：状态变更后放行重读", "增强(我独有)"),
    (19, "状态管理", "待办/挂起动作快照",
     "（无对应）",
     "⑫ PendingActionStore 前后快照差集收集 new_pending",
     "我方独有：挂起动作归集", "增强(我独有)"),
]

r = hrow + 1
for idx, (n, cat, cap, cla, mine, gap, status) in enumerate(rows):
    fill = band_fill if idx % 2 else None
    vals = [n, cat, cap, cla, mine, gap, status]
    for ci, v in enumerate(vals, start=1):
        c = s1.cell(row=r, column=ci, value=v)
        c.font = cell_font; c.alignment = wrap_top; c.border = border
        if fill: c.fill = fill
    sc = s1.cell(row=r, column=7)
    sc.fill = status_map[status]; sc.font = status_font[status]; sc.alignment = center
    s1.cell(row=r, column=1).alignment = center
    s1.cell(row=r, column=2).font = bold_cell
    r += 1

widths = [4, 12, 24, 34, 38, 40, 13]
for i, w in enumerate(widths, start=1):
    s1.column_dimensions[get_column_letter(i)].width = w
s1.freeze_panes = "A8"

# ================= Sheet 2: 未实现差距清单 =================
s2 = wb.create_sheet("未实现差距清单")
s2["A1"] = "未实现 / 部分覆盖能力清单（需补强的项）"
s2["A1"].font = title_font
s2.merge_cells("A1:E1")

g_headers = ["#", "差距能力", "Claude Code 对应环节", "我的现状", "建议落地方式"]
for i, h in enumerate(g_headers, start=1):
    c = s2.cell(row=3, column=i, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border

gaps = [
    (1, "泛型输入校验 validateInput",
     "④ validateInput() 失败返回错误 tool_result",
     "仅写操作有 precondition 断言，读/中性工具无统一 schema 校验",
     "在 _handle_call 入口加一层泛型校验：用工具声明的 input schema（JSON Schema / Pydantic）统一校验所有 call，失败即生成错误 observation 并跳过，对齐 Claude 的 ④"),
    (2, "实时进度回调 onProgress",
     "⑦ call() 内 onProgress() 实时更新 UI",
     "handler 执行期间无流式进度回调，长任务 UI 静止",
     "为 handler 增加 on_progress(callback) 钩子，在 registry.execute 与 action 执行中分阶段 emit 进度事件，前端订阅渲染；可参考 Claude 的 onProgress 回调模型"),
    (3, "多工具并发执行 runTools",
     "② StreamingToolExecutor.runTools() 并发",
     "一轮多 call 在 ③ 中逐条 _handle_call 串行",
     "对相互独立（无写依赖）的 tool_call 做并行池（asyncio.gather / ThreadPool），有写依赖的保持串行；注意并发下的 seen 去重与 ActionGate 竞争"),
    (4, "交互式权限确认 canUseTool",
     "⑤ Ask 模式弹确认，拒绝返回拒绝 tool_result",
     "仅写操作走 ActionGate.pending，读/中性工具默认放行",
     "抽象通用 can_use_tool(name, args) 拦截层：在 ④ 白名单之后、⑥ 执行之前，对所有工具提供'询问用户确认'的交互入口，pending 时挂起而非默认通过"),
    (5, "独立规则权限引擎 checkPermissions",
     "⑥ checkPermissions() 独立规则匹配",
     "权限规则与写操作耦合在 ActionGate 内",
     "抽出独立的权限规则引擎（允许/拒绝/询问三态），与 ActionGate 解耦，对读/写/中性工具统一评估，提升可配置性与可审计性"),
]

r = 4
for idx, (n, name, cla, mine, sug) in enumerate(gaps):
    fill = band_fill if idx % 2 else None
    vals = [n, name, cla, mine, sug]
    for ci, v in enumerate(vals, start=1):
        c = s2.cell(row=r, column=ci, value=v)
        c.font = cell_font; c.alignment = wrap_top; c.border = border
        if fill: c.fill = fill
    s2.cell(row=r, column=1).alignment = center
    s2.cell(row=r, column=2).font = bold_cell
    r += 1

g_widths = [4, 26, 34, 38, 50]
for i, w in enumerate(g_widths, start=1):
    s2.column_dimensions[get_column_letter(i)].width = w
s2.freeze_panes = "A4"

# ================= Sheet 3: 步骤链路对照 =================
s3 = wb.create_sheet("步骤链路对照")
s3["A1"] = "调用链步骤编号对照（Claude ①-⑩ ↔ 我的 ①-⑫）"
s3["A1"].font = title_font
s3.merge_cells("A1:C1")
for i, h in enumerate(["Claude Code 环节", "我的对应环节", "说明"], start=1):
    c = s3.cell(row=3, column=i, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
link = [
    ("① API 返回 tool_use block", "① llm.chat_turn() 返回 AgentTurn(tool_calls[])", "解析入口，我方支持多 call"),
    ("② StreamingToolExecutor.runTools()", "② run() 外层循环 / max_steps / _is_stuck", "编排层，我方多了上限与卡死检测"),
    ("③ findToolByName()", "⑥ self._tools.get(name)", "工具定位"),
    ("④ validateInput()", "⑦ tool.precondition(args) + ⑦b self._extra", "我方仅写操作有断言，缺泛型校验"),
    ("⑤ canUseTool()（Ask 确认）", "⑧ ActionGate.request(pending)（仅写）", "我方仅写操作询问，缺通用确认层"),
    ("⑥ checkPermissions()", "⑧ ActionGate 规则（耦合写）", "我方无独立权限引擎"),
    ("⑦ call() + onProgress()", "⑧ self._tools.execute → handler(**args)", "执行，我方缺 onProgress 进度回调"),
    ("⑧ ToolResult<Output>", "⑧ handler 返回 + _serialize_observation", "结果封装"),
    ("⑨ mapToolResultToToolResultBlockParam()", "⑩ _serialize_observation + ⑪ append role:tool", "结果序列化追加"),
    ("⑩ 追加消息→下一轮", "⑪ append + ⑫ PendingActionStore diff", "我方额外做挂起动作归集"),
    ("（无）", "③ EMPTY→nudge / ⑤ seen 去重 / ⑨ 清读 seen", "我方独有护栏"),
    ("（无）", "④ 护栏3 白名单", "我方独有白名单"),
]
r = 4
for idx, (a, b, note) in enumerate(link):
    fill = band_fill if idx % 2 else None
    for ci, v in enumerate([a, b, note], start=1):
        c = s3.cell(row=r, column=ci, value=v)
        c.font = cell_font; c.alignment = wrap_top; c.border = border
        if fill: c.fill = fill
    r += 1
for i, w in enumerate([34, 42, 40], start=1):
    s3.column_dimensions[get_column_letter(i)].width = w
s3.freeze_panes = "A4"

out = "/Users/zhouwentao/Desktop/manufacturing-agent/工具调用链能力对比.xlsx"
wb.save(out)
print("saved", out)
